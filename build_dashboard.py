import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

df = pd.read_excel(r'C:\Users\luker\Downloads\luke-data.xlsx')
df['Request Date'] = pd.to_datetime(df['Request Date'])

states = sorted(df['State'].unique())
techs = sorted(df['Technician'].unique())

# ── helpers ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ALT_ROW    = "EBF3FB"
WHITE      = "FFFFFF"
FONT_NAME  = "Arial"

def hdr_cell(cell, value, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
    cell.value = value
    cell.font = Font(bold=bold, name=FONT_NAME, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DATA  (Excel Table with built-in AutoFilter on every column)
# ════════════════════════════════════════════════════════════════════════════
ws_d = wb.active
ws_d.title = "Data"

# ── banner ──
ws_d.merge_cells("A1:H1")
banner = ws_d["A1"]
banner.value = "Job Data — use column dropdowns to Filter by Status · Payment · State · Invoice · Tech"
banner.font = Font(bold=True, name=FONT_NAME, color=WHITE, size=11)
banner.fill = PatternFill("solid", fgColor=MID_BLUE)
banner.alignment = Alignment(horizontal="center", vertical="center")
ws_d.row_dimensions[1].height = 28

# ── column headers (row 2) ──
headers = list(df.columns)
for ci, h in enumerate(headers, 1):
    hdr_cell(ws_d.cell(row=2, column=ci), h)
ws_d.row_dimensions[2].height = 28

# ── data rows ──
for ri, row in df.iterrows():
    er = ri + 3
    for ci, val in enumerate(row, 1):
        if hasattr(val, 'date'):
            val = val.date()
        cell = ws_d.cell(row=er, column=ci, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ri % 2 == 1:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW)
        if ci == 5:
            cell.number_format = '$#,##0'

# ── Excel Table (gives dropdown filters + sort on every column) ──
last_row = len(df) + 2
tab = Table(displayName="DataTable",
            ref=f"A2:{get_column_letter(len(headers))}{last_row}")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showRowStripes=True, showColumnStripes=False)
ws_d.add_table(tab)

apply_border(ws_d, 2, last_row, 1, len(headers))

# ── column widths ──
for ci, w in enumerate([12, 38, 8, 14, 15, 14, 10, 12], 1):
    ws_d.column_dimensions[get_column_letter(ci)].width = w
ws_d.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — HEAT MAP  (State × Tech, total Invoice $, color scale)
# ════════════════════════════════════════════════════════════════════════════
ws_hm = wb.create_sheet("Heat Map")

# title
ws_hm.merge_cells("A1:F1")
t = ws_hm["A1"]
t.value = "Revenue Heat Map — Invoice ($) by State & Technician"
t.font = Font(bold=True, name=FONT_NAME, size=13, color=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_hm.row_dimensions[1].height = 34

ws_hm.merge_cells("A2:F2")
n = ws_hm["A2"]
n.value = "Darker green = higher revenue.  Use Data sheet column dropdowns to filter by Status, Payment, State, Invoice, or Tech."
n.font = Font(italic=True, name=FONT_NAME, size=9, color="595959")
n.alignment = Alignment(horizontal="center")
ws_hm.row_dimensions[2].height = 18

# header row (row 3)
HR = 3
hdr_cell(ws_hm.cell(row=HR, column=1), "State")
total_col = len(techs) + 2
for ci, tech in enumerate(techs, 2):
    hdr_cell(ws_hm.cell(row=HR, column=ci), tech)
hdr_cell(ws_hm.cell(row=HR, column=total_col), "Row Total")
ws_hm.row_dimensions[HR].height = 24

# data rows
DS = HR + 1
for ri, state in enumerate(states):
    er = DS + ri
    lc = ws_hm.cell(row=er, column=1, value=state)
    lc.font = Font(bold=True, name=FONT_NAME)
    lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    ws_hm.row_dimensions[er].height = 22

    for ci, tech in enumerate(techs, 2):
        cell = ws_hm.cell(
            row=er, column=ci,
            value=f'=SUMIFS(DataTable[Invoice ($)],DataTable[State],"{state}",DataTable[Technician],"{tech}")')
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_NAME)

    first_tech_col = get_column_letter(2)
    last_tech_col  = get_column_letter(len(techs) + 1)
    tc = ws_hm.cell(row=er, column=total_col,
                    value=f"=SUM({first_tech_col}{er}:{last_tech_col}{er})")
    tc.number_format = '$#,##0'
    tc.font = Font(bold=True, name=FONT_NAME)
    tc.alignment = Alignment(horizontal="center", vertical="center")

DE = DS + len(states) - 1

# totals row
TR = DE + 1
hdr_cell(ws_hm.cell(row=TR, column=1), "Col Total")
ws_hm.row_dimensions[TR].height = 24
for ci in range(2, total_col + 1):
    cl = get_column_letter(ci)
    cell = ws_hm.cell(row=TR, column=ci, value=f"=SUM({cl}{DS}:{cl}{DE})")
    cell.number_format = '$#,##0'
    cell.font = Font(bold=True, name=FONT_NAME, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# color scale on the data cells (not total col / row)
hm_range = f"B{DS}:{get_column_letter(len(techs)+1)}{DE}"
ws_hm.conditional_formatting.add(hm_range, ColorScaleRule(
    start_type='num',       start_value=0,  start_color='FFFFFF',
    mid_type='percentile',  mid_value=50,   mid_color='A8D08D',
    end_type='max',                         end_color='375623'))

apply_border(ws_hm, HR, TR, 1, total_col)

ws_hm.column_dimensions['A'].width = 10
for ci in range(2, total_col + 1):
    ws_hm.column_dimensions[get_column_letter(ci)].width = 16
ws_hm.freeze_panes = f"B{DS}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BY TECH  (summary + state breakdown, with Invoice sort note)
# ════════════════════════════════════════════════════════════════════════════
ws_t = wb.create_sheet("By Tech")

ws_t.merge_cells("A1:G1")
tt = ws_t["A1"]
tt.value = "Technician Performance Dashboard"
tt.font = Font(bold=True, name=FONT_NAME, size=13, color=DARK_BLUE)
tt.alignment = Alignment(horizontal="center", vertical="center")
ws_t.row_dimensions[1].height = 34

# ── Section A: Overall summary ──
ws_t.merge_cells("A3:G3")
sa = ws_t["A3"]
sa.value = "OVERALL SUMMARY"
sa.font = Font(bold=True, name=FONT_NAME, size=10, color=WHITE)
sa.fill = PatternFill("solid", fgColor=MID_BLUE)
sa.alignment = Alignment(horizontal="center")
ws_t.row_dimensions[3].height = 22

t_hdrs = ["Technician", "Total Jobs", "Total Revenue ($)",
          "Avg Invoice ($)", "Closed Jobs", "Open Jobs", "Unpaid Revenue ($)"]
for ci, h in enumerate(t_hdrs, 1):
    hdr_cell(ws_t.cell(row=4, column=ci), h, size=9)
ws_t.row_dimensions[4].height = 26

for ri, tech in enumerate(techs, 5):
    ws_t.cell(row=ri, column=1, value=tech).font = Font(bold=True, name=FONT_NAME)
    ws_t.cell(row=ri, column=1).alignment = Alignment(horizontal="center", vertical="center")

    formulas = [
        f'=COUNTIF(DataTable[Technician],"{tech}")',
        f'=SUMIF(DataTable[Technician],"{tech}",DataTable[Invoice ($)])',
        f'=IFERROR(C{ri}/B{ri},0)',
        f'=COUNTIFS(DataTable[Technician],"{tech}",DataTable[Status],"Closed")',
        f'=COUNTIFS(DataTable[Technician],"{tech}",DataTable[Status],"Open")',
        f'=SUMIFS(DataTable[Invoice ($)],DataTable[Technician],"{tech}",DataTable[Payment],"Unpaid")',
    ]
    fmts = [None, '$#,##0', '$#,##0', None, None, '$#,##0']
    for ci, (formula, fmt) in enumerate(zip(formulas, fmts), 2):
        cell = ws_t.cell(row=ri, column=ci, value=formula)
        cell.font = Font(name=FONT_NAME)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt
    ws_t.row_dimensions[ri].height = 20

# totals
TR2 = 5 + len(techs)
hdr_cell(ws_t.cell(row=TR2, column=1), "TOTAL", size=9)
for ci in range(2, 8):
    cl = get_column_letter(ci)
    cell = ws_t.cell(row=TR2, column=ci, value=f"=SUM({cl}5:{cl}{TR2-1})")
    cell.font = Font(bold=True, name=FONT_NAME, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center")
    if ci in [3, 4, 7]:
        cell.number_format = '$#,##0'
ws_t.row_dimensions[TR2].height = 22

apply_border(ws_t, 4, TR2, 1, 7)

# ── Section B: Revenue by Tech × State (heatmap grid) ──
S2 = TR2 + 2
ws_t.merge_cells(f"A{S2}:{get_column_letter(1+len(states))}{S2}")
sb = ws_t[f"A{S2}"]
sb.value = "REVENUE BY TECHNICIAN & STATE ($)  —  Darker = higher revenue"
sb.font = Font(bold=True, name=FONT_NAME, size=10, color=WHITE)
sb.fill = PatternFill("solid", fgColor=MID_BLUE)
sb.alignment = Alignment(horizontal="center")
ws_t.row_dimensions[S2].height = 22

HR2 = S2 + 1
hdr_cell(ws_t.cell(row=HR2, column=1), "Technician")
for ci, state in enumerate(states, 2):
    hdr_cell(ws_t.cell(row=HR2, column=ci), state)
ws_t.row_dimensions[HR2].height = 24

DS2 = HR2 + 1
for ri, tech in enumerate(techs, DS2):
    ws_t.cell(row=ri, column=1, value=tech).font = Font(bold=True, name=FONT_NAME)
    ws_t.cell(row=ri, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_t.row_dimensions[ri].height = 20
    for ci, state in enumerate(states, 2):
        cell = ws_t.cell(
            row=ri, column=ci,
            value=f'=SUMIFS(DataTable[Invoice ($)],DataTable[Technician],"{tech}",DataTable[State],"{state}")')
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT_NAME)

DE2 = DS2 + len(techs) - 1

# color scale
grid2 = f"B{DS2}:{get_column_letter(1+len(states))}{DE2}"
ws_t.conditional_formatting.add(grid2, ColorScaleRule(
    start_type='num',      start_value=0, start_color='FFFFFF',
    mid_type='percentile', mid_value=50,  mid_color='A8D08D',
    end_type='max',                       end_color='375623'))

apply_border(ws_t, HR2, DE2, 1, 1 + len(states))

# ── Section C: Invoice Sort Guide ──
S3 = DE2 + 2
ws_t.merge_cells(f"A{S3}:G{S3}")
sc = ws_t[f"A{S3}"]
sc.value = "SORT BY INVOICE — Go to the Data sheet and click the Invoice ($) column dropdown → Sort Largest to Smallest (or A→Z)"
sc.font = Font(italic=True, name=FONT_NAME, size=9, color="595959")
sc.fill = PatternFill("solid", fgColor="FFF2CC")
sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_t.row_dimensions[S3].height = 30

# column widths
ws_t.column_dimensions['A'].width = 15
for ci in range(2, max(8, 2 + len(states))):
    ws_t.column_dimensions[get_column_letter(ci)].width = 16

ws_t.freeze_panes = "B5"

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out = r'C:\Users\luker\Downloads\luke-dashboard.xlsx'
wb.save(out)
print(f"Saved: {out}")
